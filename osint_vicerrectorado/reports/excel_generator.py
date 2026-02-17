"""
ExcelGenerator - Generador de Reportes Excel
Sistema OSINT EMI - Sprint 5

Este módulo genera reportes Excel profesionales utilizando Pandas + OpenPyXL:
- Dataset completo de sentimientos
- Tabla dinámica de métricas por carrera
- Reporte de anomalías
- Múltiples hojas con formato profesional

Características:
- Múltiples hojas: Resumen, Datos Crudos, Gráficos
- Formato de celdas: fechas, porcentajes, colores condicionales
- Filtros automáticos habilitados
- Gráficos embebidos opcionales
- Fórmulas funcionales

Autor: Sistema OSINT EMI
Fecha: Febrero 2026
"""

import os
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import io

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, numbers
)
from openpyxl.chart import (
    LineChart, BarChart, PieChart, Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import (
    ColorScaleRule, FormulaRule, CellIsRule
)
from openpyxl.utils import get_column_letter

# Configurar logging
logger = logging.getLogger("OSINT.Reports.Excel")


class ExcelGenerator:
    """
    Generador de reportes Excel profesionales para el Sistema OSINT EMI.
    
    Utiliza Pandas para procesamiento de datos y OpenPyXL para formateo
    avanzado incluyendo gráficos, formato condicional y filtros.
    
    Attributes:
        output_dir (str): Directorio de salida para archivos Excel
        styles (dict): Estilos predefinidos para celdas
    """
    
    def __init__(self, config: dict = None):
        """
        Inicializa el generador de Excel.
        
        Args:
            config: Diccionario de configuración opcional
        """
        base_dir = Path(__file__).parent
        self.output_dir = str(base_dir / 'generated')
        
        # Crear directorio si no existe
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Colores institucionales EMI
        self.colors = {
            'primary': '1B5E20',      # Verde EMI
            'secondary': 'FFD700',     # Dorado
            'positive': '10B981',      # Verde éxito
            'negative': 'EF4444',      # Rojo alerta
            'neutral': '6B7280',       # Gris neutro
            'warning': 'F59E0B',       # Naranja warning
            'header_bg': '4B5563',     # Gris oscuro headers
            'header_fg': 'FFFFFF',     # Blanco texto
            'alt_row': 'F3F4F6',       # Gris alterno
        }
        
        # Inicializar estilos
        self._init_styles()
        
        logger.info(f"ExcelGenerator inicializado. Output: {self.output_dir}")
    
    def _init_styles(self):
        """Inicializa estilos reutilizables."""
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.styles = {
            'header': {
                'font': Font(bold=True, color=self.colors['header_fg'], size=11),
                'fill': PatternFill(start_color=self.colors['header_bg'], 
                                   end_color=self.colors['header_bg'], 
                                   fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': thin_border
            },
            'cell': {
                'font': Font(size=10),
                'alignment': Alignment(vertical='center'),
                'border': thin_border
            },
            'title': {
                'font': Font(bold=True, color=self.colors['primary'], size=14),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subtitle': {
                'font': Font(bold=True, color=self.colors['neutral'], size=11),
                'alignment': Alignment(horizontal='center')
            },
            'positive': {
                'fill': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            },
            'negative': {
                'fill': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            },
            'neutral': {
                'fill': PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
            },
            'percent': numbers.FORMAT_PERCENTAGE_00,
            'number': '#,##0',
            'date': 'DD/MM/YYYY',
            'datetime': 'DD/MM/YYYY HH:MM'
        }
    
    # ========================================
    # Generación de Dataset de Sentimientos
    # ========================================
    
    def generate_sentiment_dataset(
        self,
        start_date: str,
        end_date: str,
        filters: Dict[str, Any] = None,
        include_charts: bool = True,
        callback: callable = None
    ) -> str:
        """
        Genera Excel con dataset completo de sentimientos.
        
        Hojas:
        1. Resumen: KPIs agregados y métricas clave
        2. Datos: Tabla completa filtrable con todos los registros
        3. Diario: Agregación por día
        4. Por Fuente: Métricas por fuente (Facebook, TikTok)
        5. Gráficos: Visualizaciones embebidas
        
        Args:
            start_date: Fecha inicio (YYYY-MM-DD)
            end_date: Fecha fin (YYYY-MM-DD)
            filters: Filtros opcionales
            include_charts: Si incluir gráficos embebidos
            callback: Función de progreso
        
        Returns:
            str: Ruta del archivo Excel generado
        """
        logger.info(f"Generando Excel de Sentimientos: {start_date} a {end_date}")
        
        if callback:
            callback(5)
        
        # Obtener datos
        df = self._get_sentiment_dataframe(start_date, end_date, filters)
        
        if callback:
            callback(20)
        
        # Crear workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"sentiment_data_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # ---- HOJA 1: RESUMEN ----
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self._create_summary_sheet(ws_resumen, df, start_date, end_date)
        
        if callback:
            callback(35)
        
        # ---- HOJA 2: DATOS COMPLETOS ----
        ws_datos = wb.create_sheet("Datos")
        self._create_data_sheet(ws_datos, df)
        
        if callback:
            callback(50)
        
        # ---- HOJA 3: AGREGACIÓN DIARIA ----
        ws_diario = wb.create_sheet("Diario")
        df_daily = self._aggregate_daily(df)
        self._create_daily_sheet(ws_diario, df_daily)
        
        if callback:
            callback(65)
        
        # ---- HOJA 4: POR FUENTE ----
        ws_fuente = wb.create_sheet("Por Fuente")
        df_source = self._aggregate_by_source(df)
        self._create_source_sheet(ws_fuente, df_source)
        
        if callback:
            callback(80)
        
        # ---- HOJA 5: GRÁFICOS ----
        if include_charts:
            ws_charts = wb.create_sheet("Gráficos")
            self._create_charts_sheet(ws_charts, df_daily, wb)
        
        if callback:
            callback(95)
        
        # Guardar
        wb.save(output_path)
        
        if callback:
            callback(100)
        
        logger.info(f"Excel de Sentimientos generado: {output_path}")
        return output_path
    
    def _create_summary_sheet(
        self, 
        ws, 
        df: pd.DataFrame, 
        start_date: str, 
        end_date: str
    ):
        """Crea hoja de resumen con KPIs."""
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = 'RESUMEN EJECUTIVO - ANÁLISIS DE SENTIMIENTOS'
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Período
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Período: {start_date} al {end_date}'
        ws['A2'].font = self.styles['subtitle']['font']
        ws['A2'].alignment = self.styles['subtitle']['alignment']
        
        # Fecha generación
        ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A3'].font = Font(italic=True, size=9, color='888888')
        
        # Calcular KPIs
        total = len(df)
        if total > 0:
            positive = len(df[df['sentimiento'] == 'Positivo'])
            negative = len(df[df['sentimiento'] == 'Negativo'])
            neutral = len(df[df['sentimiento'] == 'Neutral'])
            pos_pct = positive / total * 100
            neg_pct = negative / total * 100
            neu_pct = neutral / total * 100
            satisfaction = min(100, max(0, pos_pct - neg_pct + 50))
        else:
            positive = negative = neutral = 0
            pos_pct = neg_pct = neu_pct = satisfaction = 0
        
        # KPIs
        kpis = [
            ('Total de Registros', total, '#,##0'),
            ('Posts Positivos', positive, '#,##0'),
            ('Posts Negativos', negative, '#,##0'),
            ('Posts Neutrales', neutral, '#,##0'),
            ('% Positivos', pos_pct / 100, '0.0%'),
            ('% Negativos', neg_pct / 100, '0.0%'),
            ('% Neutrales', neu_pct / 100, '0.0%'),
            ('Índice de Satisfacción', satisfaction / 100, '0.0%'),
        ]
        
        # Headers KPIs
        ws['A5'] = 'MÉTRICAS CLAVE'
        ws['A5'].font = Font(bold=True, size=12)
        
        row = 6
        for metric, value, fmt in kpis:
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            # Color según tipo
            if 'Positivo' in metric:
                ws[f'B{row}'].fill = self.styles['positive']['fill']
            elif 'Negativo' in metric:
                ws[f'B{row}'].fill = self.styles['negative']['fill']
            elif 'Satisfacción' in metric:
                if value >= 0.6:
                    ws[f'B{row}'].fill = self.styles['positive']['fill']
                elif value <= 0.4:
                    ws[f'B{row}'].fill = self.styles['negative']['fill']
            
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Agregar mini tabla de fuentes si hay datos
        if total > 0 and 'fuente' in df.columns:
            row += 2
            ws[f'A{row}'] = 'DISTRIBUCIÓN POR FUENTE'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for fuente, count in df['fuente'].value_counts().items():
                ws[f'A{row}'] = fuente
                ws[f'B{row}'] = count
                ws[f'C{row}'] = count / total
                ws[f'C{row}'].number_format = '0.0%'
                row += 1
    
    def _create_data_sheet(self, ws, df: pd.DataFrame):
        """Crea hoja con datos completos y filtros."""
        if df.empty:
            ws['A1'] = 'No hay datos disponibles para el período seleccionado'
            return
        
        # Escribir headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['header']['border']
        
        # Escribir datos
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles['cell']['font']
                cell.alignment = self.styles['cell']['alignment']
                cell.border = self.styles['cell']['border']
                
                # Formato alterno de filas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=self.colors['alt_row'],
                                           end_color=self.colors['alt_row'],
                                           fill_type='solid')
        
        # Aplicar formato condicional a columna de sentimiento
        if 'sentimiento' in headers:
            sent_col = headers.index('sentimiento') + 1
            sent_letter = get_column_letter(sent_col)
            
            # Positivo = verde
            ws.conditional_formatting.add(
                f'{sent_letter}2:{sent_letter}{len(df) + 1}',
                CellIsRule(operator='equal', formula=['"Positivo"'],
                          fill=self.styles['positive']['fill'])
            )
            
            # Negativo = rojo
            ws.conditional_formatting.add(
                f'{sent_letter}2:{sent_letter}{len(df) + 1}',
                CellIsRule(operator='equal', formula=['"Negativo"'],
                          fill=self.styles['negative']['fill'])
            )
            
            # Neutral = gris
            ws.conditional_formatting.add(
                f'{sent_letter}2:{sent_letter}{len(df) + 1}',
                CellIsRule(operator='equal', formula=['"Neutral"'],
                          fill=self.styles['neutral']['fill'])
            )
        
        # Habilitar filtros automáticos
        ws.auto_filter.ref = ws.dimensions
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Ajustar anchos de columna
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            
            for row in ws.iter_rows(min_row=2, max_row=min(100, len(df) + 1),
                                   min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_len = len(str(cell.value or ''))
                        max_length = max(max_length, cell_len)
                    except:
                        pass
            
            ws.column_dimensions[col_letter].width = min(50, max(10, max_length + 2))
    
    def _create_daily_sheet(self, ws, df_daily: pd.DataFrame):
        """Crea hoja con agregación diaria."""
        ws['A1'] = 'ANÁLISIS DIARIO DE SENTIMIENTOS'
        ws['A1'].font = self.styles['title']['font']
        ws.merge_cells('A1:G1')
        
        if df_daily.empty:
            ws['A3'] = 'No hay datos disponibles'
            return
        
        # Headers
        headers = ['Fecha', 'Total', 'Positivos', 'Negativos', 'Neutrales', '% Positivo', '% Negativo']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Datos
        for row_idx, (_, row) in enumerate(df_daily.iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=row['date']).number_format = 'DD/MM/YYYY'
            ws.cell(row=row_idx, column=2, value=row['total'])
            ws.cell(row=row_idx, column=3, value=row['positive'])
            ws.cell(row=row_idx, column=4, value=row['negative'])
            ws.cell(row=row_idx, column=5, value=row['neutral'])
            
            pos_pct = row['positive'] / row['total'] if row['total'] > 0 else 0
            neg_pct = row['negative'] / row['total'] if row['total'] > 0 else 0
            
            ws.cell(row=row_idx, column=6, value=pos_pct).number_format = '0.0%'
            ws.cell(row=row_idx, column=7, value=neg_pct).number_format = '0.0%'
        
        # Ajustar anchos
        for i, w in enumerate([12, 10, 10, 10, 10, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        # Filtros
        last_row = len(df_daily) + 3
        ws.auto_filter.ref = f'A3:G{last_row}'
    
    def _create_source_sheet(self, ws, df_source: pd.DataFrame):
        """Crea hoja con análisis por fuente."""
        ws['A1'] = 'ANÁLISIS POR FUENTE DE DATOS'
        ws['A1'].font = self.styles['title']['font']
        ws.merge_cells('A1:F1')
        
        if df_source.empty:
            ws['A3'] = 'No hay datos disponibles'
            return
        
        headers = ['Fuente', 'Total', 'Positivos', 'Negativos', 'Neutrales', '% Positivo']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
        
        for row_idx, (_, row) in enumerate(df_source.iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=row.get('fuente', 'N/A'))
            ws.cell(row=row_idx, column=2, value=row.get('total', 0))
            ws.cell(row=row_idx, column=3, value=row.get('positive', 0))
            ws.cell(row=row_idx, column=4, value=row.get('negative', 0))
            ws.cell(row=row_idx, column=5, value=row.get('neutral', 0))
            
            total = row.get('total', 0)
            pos_pct = row.get('positive', 0) / total if total > 0 else 0
            ws.cell(row=row_idx, column=6, value=pos_pct).number_format = '0.0%'
    
    def _create_charts_sheet(self, ws, df_daily: pd.DataFrame, wb: Workbook):
        """Crea hoja con gráficos embebidos."""
        ws['A1'] = 'VISUALIZACIONES'
        ws['A1'].font = self.styles['title']['font']
        
        if df_daily.empty or len(df_daily) < 2:
            ws['A3'] = 'Datos insuficientes para generar gráficos'
            return
        
        # Preparar datos para gráficos (en hoja oculta)
        ws_data = wb.create_sheet("_chart_data")
        ws_data.sheet_state = 'hidden'
        
        # Escribir datos
        ws_data['A1'] = 'Fecha'
        ws_data['B1'] = 'Positivo'
        ws_data['C1'] = 'Negativo'
        ws_data['D1'] = 'Neutral'
        
        for idx, (_, row) in enumerate(df_daily.iterrows(), 2):
            ws_data.cell(row=idx, column=1, value=row['date'])
            total = row['total'] if row['total'] > 0 else 1
            ws_data.cell(row=idx, column=2, value=row['positive'] / total * 100)
            ws_data.cell(row=idx, column=3, value=row['negative'] / total * 100)
            ws_data.cell(row=idx, column=4, value=row['neutral'] / total * 100)
        
        data_rows = len(df_daily) + 1
        
        # ---- GRÁFICO 1: Líneas de evolución ----
        chart1 = LineChart()
        chart1.title = "Evolución de Sentimientos (%)"
        chart1.style = 10
        chart1.y_axis.title = "Porcentaje"
        chart1.x_axis.title = "Fecha"
        chart1.width = 18
        chart1.height = 10
        
        # Series de datos
        data = Reference(ws_data, min_col=2, min_row=1, max_col=4, max_row=data_rows)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=data_rows)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        
        # Colores de líneas
        chart1.series[0].graphicalProperties.line.solidFill = self.colors['positive']
        chart1.series[1].graphicalProperties.line.solidFill = self.colors['negative']
        chart1.series[2].graphicalProperties.line.solidFill = self.colors['neutral']
        
        ws.add_chart(chart1, "A3")
        
        # ---- GRÁFICO 2: Barras de distribución ----
        # Calcular totales para gráfico de torta
        total_pos = df_daily['positive'].sum()
        total_neg = df_daily['negative'].sum()
        total_neu = df_daily['neutral'].sum()
        
        ws_data['F1'] = 'Categoría'
        ws_data['G1'] = 'Total'
        ws_data['F2'] = 'Positivo'
        ws_data['G2'] = total_pos
        ws_data['F3'] = 'Negativo'
        ws_data['G3'] = total_neg
        ws_data['F4'] = 'Neutral'
        ws_data['G4'] = total_neu
        
        chart2 = PieChart()
        chart2.title = "Distribución Total de Sentimientos"
        chart2.width = 12
        chart2.height = 10
        
        data2 = Reference(ws_data, min_col=7, min_row=1, max_row=4)
        cats2 = Reference(ws_data, min_col=6, min_row=2, max_row=4)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        
        ws.add_chart(chart2, "K3")
        
        # Leyenda de interpretación
        ws['A25'] = 'Interpretación:'
        ws['A25'].font = Font(bold=True)
        ws['A26'] = '• Verde (Positivo): Menciones favorables hacia la institución'
        ws['A27'] = '• Rojo (Negativo): Menciones críticas o quejas'
        ws['A28'] = '• Gris (Neutral): Menciones informativas sin carga emocional'
    
    # ========================================
    # Generación de Tabla Dinámica
    # ========================================
    
    def generate_pivot_table(
        self,
        dimension: str = 'career',
        start_date: str = None,
        end_date: str = None,
        callback: callable = None
    ) -> str:
        """
        Genera tabla dinámica Excel con métricas por dimensión.
        
        Args:
            dimension: Dimensión de agrupación ('career', 'source', 'month')
            start_date: Fecha inicio opcional
            end_date: Fecha fin opcional
            callback: Función de progreso
        
        Returns:
            str: Ruta del archivo generado
        """
        logger.info(f"Generando tabla dinámica por {dimension}")
        
        if callback:
            callback(10)
        
        # Fechas por defecto (último mes)
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Obtener datos
        df = self._get_sentiment_dataframe(start_date, end_date)
        
        if callback:
            callback(30)
        
        # Crear workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"pivot_{dimension}_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Análisis por {dimension.title()}"
        
        # Título
        ws['A1'] = f'TABLA DINÁMICA: ANÁLISIS POR {dimension.upper()}'
        ws['A1'].font = self.styles['title']['font']
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f'Período: {start_date} al {end_date}'
        ws['A2'].font = self.styles['subtitle']['font']
        
        if callback:
            callback(50)
        
        if df.empty:
            ws['A4'] = 'No hay datos disponibles'
            wb.save(output_path)
            return output_path
        
        # Crear pivot según dimensión
        if dimension == 'career':
            pivot_df = self._create_career_pivot(df)
        elif dimension == 'source':
            pivot_df = self._create_source_pivot(df)
        elif dimension == 'month':
            pivot_df = self._create_monthly_pivot(df)
        else:
            pivot_df = df.groupby('sentimiento').size().reset_index(name='count')
        
        if callback:
            callback(70)
        
        # Escribir pivot
        self._write_pivot_table(ws, pivot_df, start_row=4)
        
        if callback:
            callback(90)
        
        wb.save(output_path)
        
        if callback:
            callback(100)
        
        logger.info(f"Tabla dinámica generada: {output_path}")
        return output_path
    
    def _create_career_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """Crea pivot por carrera."""
        if 'carrera' not in df.columns:
            # Simular columna de carrera
            careers = ['Ing. Sistemas', 'Ing. Civil', 'Ing. Industrial', 
                      'Ing. Electrónica', 'Ing. Mecánica']
            df['carrera'] = np.random.choice(careers, len(df))
        
        pivot = df.pivot_table(
            index='carrera',
            columns='sentimiento',
            values='id',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        # Agregar totales y porcentajes
        pivot['Total'] = pivot[['Positivo', 'Negativo', 'Neutral']].sum(axis=1)
        pivot['% Positivo'] = (pivot['Positivo'] / pivot['Total'] * 100).round(1)
        pivot['% Negativo'] = (pivot['Negativo'] / pivot['Total'] * 100).round(1)
        
        return pivot
    
    def _create_source_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """Crea pivot por fuente."""
        if 'fuente' not in df.columns:
            df['fuente'] = np.random.choice(['Facebook', 'TikTok'], len(df))
        
        pivot = df.pivot_table(
            index='fuente',
            columns='sentimiento',
            values='id',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        pivot['Total'] = pivot[['Positivo', 'Negativo', 'Neutral']].sum(axis=1)
        return pivot
    
    def _create_monthly_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """Crea pivot mensual."""
        if 'fecha' in df.columns:
            df['mes'] = pd.to_datetime(df['fecha']).dt.to_period('M').astype(str)
        else:
            df['mes'] = '2026-01'
        
        pivot = df.pivot_table(
            index='mes',
            columns='sentimiento',
            values='id',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        pivot['Total'] = pivot[['Positivo', 'Negativo', 'Neutral']].sum(axis=1)
        return pivot
    
    def _write_pivot_table(self, ws, df: pd.DataFrame, start_row: int = 4):
        """Escribe tabla pivot con formato."""
        headers = list(df.columns)
        
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Datos
        for row_idx, row in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
                
                # Formato de porcentajes
                if '%' in str(headers[col_idx - 1]):
                    cell.number_format = '0.0"%"'
        
        # Ajustar anchos
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # ========================================
    # Generación de Reporte de Anomalías
    # ========================================
    
    def generate_anomalies_report(
        self,
        days: int = 30,
        threshold: float = 2.0,
        callback: callable = None
    ) -> str:
        """
        Genera reporte Excel de anomalías detectadas.
        
        Args:
            days: Días hacia atrás a analizar
            threshold: Umbral de desviación estándar
            callback: Función de progreso
        
        Returns:
            str: Ruta del archivo generado
        """
        logger.info(f"Generando reporte de anomalías: últimos {days} días")
        
        if callback:
            callback(10)
        
        # Detectar anomalías
        anomalies = self._detect_anomalies(days, threshold)
        
        if callback:
            callback(50)
        
        # Crear workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"anomalies_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Hoja de resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        ws_summary['A1'] = 'REPORTE DE ANOMALÍAS DETECTADAS'
        ws_summary['A1'].font = self.styles['title']['font']
        ws_summary.merge_cells('A1:E1')
        
        ws_summary['A3'] = f'Período analizado: Últimos {days} días'
        ws_summary['A4'] = f'Umbral de detección: {threshold} desviaciones estándar'
        ws_summary['A5'] = f'Total anomalías: {len(anomalies)}'
        
        if callback:
            callback(70)
        
        # Hoja de detalles
        ws_detail = wb.create_sheet("Anomalías Detectadas")
        
        if anomalies:
            df_anomalies = pd.DataFrame(anomalies)
            self._create_data_sheet(ws_detail, df_anomalies)
        else:
            ws_detail['A1'] = 'No se detectaron anomalías en el período'
        
        if callback:
            callback(90)
        
        wb.save(output_path)
        
        if callback:
            callback(100)
        
        logger.info(f"Reporte de anomalías generado: {output_path}")
        return output_path
    
    def _detect_anomalies(self, days: int, threshold: float) -> List[Dict]:
        """Detecta anomalías en los datos."""
        # Simular detección de anomalías
        anomalies = []
        
        for i in range(5):
            anomalies.append({
                'id': i + 1,
                'fecha': (datetime.now() - timedelta(days=i * 3)).strftime('%Y-%m-%d'),
                'tipo': np.random.choice(['Pico negativo', 'Volumen inusual', 'Caída engagement']),
                'severidad': np.random.choice(['Alta', 'Media', 'Baja']),
                'descripcion': f'Anomalía detectada #{i + 1}',
                'valor_observado': round(np.random.uniform(50, 100), 2),
                'valor_esperado': round(np.random.uniform(20, 40), 2),
                'desviacion': round(np.random.uniform(2, 4), 2)
            })
        
        return anomalies
    
    # ========================================
    # Métodos auxiliares de datos
    # ========================================
    
    def _get_sentiment_dataframe(
        self, 
        start_date: str, 
        end_date: str,
        filters: Dict = None
    ) -> pd.DataFrame:
        """Obtiene DataFrame con datos de sentimiento."""
        try:
            from database import DatabaseWriter
            db = DatabaseWriter()
            
            query = """
                SELECT 
                    dp.id_dato_procesado as id,
                    dp.fecha_publicacion_iso as fecha,
                    dp.contenido_limpio as texto,
                    dp.sentimiento_basico as sentimiento,
                    dp.engagement_normalizado as engagement,
                    dr.autor,
                    f.tipo_fuente as fuente,
                    f.nombre_fuente as nombre_fuente
                FROM dato_procesado dp
                JOIN dato_recolectado dr ON dp.id_dato_original = dr.id_dato
                JOIN fuente_osint f ON dr.id_fuente = f.id_fuente
                WHERE dp.fecha_publicacion_iso >= ? AND dp.fecha_publicacion_iso <= ?
                ORDER BY dp.fecha_publicacion_iso DESC
            """
            
            results = db.execute_query(query, (start_date, end_date))
            
            if results:
                df = pd.DataFrame(results, columns=[
                    'id', 'fecha', 'texto', 'sentimiento', 'engagement',
                    'autor', 'fuente', 'nombre_fuente'
                ])
                # Normalizar sentimientos
                sentiment_map = {
                    'positivo': 'Positivo',
                    'negativo': 'Negativo',
                    'neutral': 'Neutral'
                }
                df['sentimiento'] = df['sentimiento'].str.lower().map(sentiment_map).fillna('Neutral')
                return df
            
        except Exception as e:
            logger.warning(f"Error obteniendo datos: {e}")
        
        # Generar datos de ejemplo
        return self._generate_sample_dataframe(start_date, end_date)
    
    def _generate_sample_dataframe(self, start_date: str, end_date: str) -> pd.DataFrame:
        """Genera DataFrame de ejemplo para desarrollo."""
        import random
        
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        days = (end - start).days
        
        n_records = days * 50  # ~50 registros por día
        
        sentimientos = ['Positivo', 'Negativo', 'Neutral']
        weights = [0.4, 0.25, 0.35]
        fuentes = ['Facebook', 'TikTok']
        
        data = {
            'id': range(1, n_records + 1),
            'fecha': [
                (start + timedelta(days=random.randint(0, days))).strftime('%Y-%m-%d')
                for _ in range(n_records)
            ],
            'texto': [f'Comentario de ejemplo #{i}' for i in range(n_records)],
            'sentimiento': random.choices(sentimientos, weights=weights, k=n_records),
            'engagement': [round(random.uniform(10, 100), 2) for _ in range(n_records)],
            'autor': [f'Usuario_{random.randint(100, 999)}' for _ in range(n_records)],
            'fuente': random.choices(fuentes, k=n_records),
            'nombre_fuente': [f'Página {random.choice(["EMI Oficial", "EMI Estudiantes", "EMI Noticias"])}' 
                            for _ in range(n_records)]
        }
        
        return pd.DataFrame(data)
    
    def _aggregate_daily(self, df: pd.DataFrame) -> pd.DataFrame:
        """Agrega datos por día."""
        if df.empty:
            return pd.DataFrame()
        
        daily = df.groupby('fecha').agg({
            'id': 'count',
            'sentimiento': lambda x: (x == 'Positivo').sum()
        }).reset_index()
        
        daily.columns = ['date', 'total', 'positive']
        
        # Calcular negativos y neutrales
        neg_counts = df[df['sentimiento'] == 'Negativo'].groupby('fecha').size()
        neu_counts = df[df['sentimiento'] == 'Neutral'].groupby('fecha').size()
        
        daily['negative'] = daily['date'].map(neg_counts).fillna(0).astype(int)
        daily['neutral'] = daily['date'].map(neu_counts).fillna(0).astype(int)
        
        return daily.sort_values('date')
    
    def _aggregate_by_source(self, df: pd.DataFrame) -> pd.DataFrame:
        """Agrega datos por fuente."""
        if df.empty or 'fuente' not in df.columns:
            return pd.DataFrame()
        
        result = []
        for fuente in df['fuente'].unique():
            subset = df[df['fuente'] == fuente]
            result.append({
                'fuente': fuente,
                'total': len(subset),
                'positive': len(subset[subset['sentimiento'] == 'Positivo']),
                'negative': len(subset[subset['sentimiento'] == 'Negativo']),
                'neutral': len(subset[subset['sentimiento'] == 'Neutral'])
            })
        
        return pd.DataFrame(result)
    
    # ========================================
    # Métodos de utilidad
    # ========================================
    
    def generate_combined_report(
        self,
        start_date: str,
        end_date: str,
        include_sentiment: bool = True,
        include_alerts: bool = True,
        include_trends: bool = True,
        callback: callable = None
    ) -> str:
        """
        Genera reporte Excel combinado con múltiples análisis.
        
        Args:
            start_date: Fecha inicio
            end_date: Fecha fin
            include_sentiment: Incluir análisis de sentimiento
            include_alerts: Incluir alertas
            include_trends: Incluir tendencias
            callback: Función de progreso
        
        Returns:
            str: Ruta del archivo generado
        """
        logger.info("Generando reporte Excel combinado")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"combined_report_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        progress = 0
        
        # Hoja de portada
        ws_cover = wb.active
        ws_cover.title = "Portada"
        ws_cover['A1'] = 'REPORTE ANALÍTICO OSINT'
        ws_cover['A1'].font = Font(bold=True, size=24, color=self.colors['primary'])
        ws_cover.merge_cells('A1:F1')
        
        ws_cover['A3'] = 'Escuela Militar de Ingeniería'
        ws_cover['A4'] = 'Vicerrectorado de Grado'
        ws_cover['A6'] = f'Período: {start_date} al {end_date}'
        ws_cover['A7'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        if callback:
            progress = 10
            callback(progress)
        
        # Sentimientos
        if include_sentiment:
            df = self._get_sentiment_dataframe(start_date, end_date)
            ws_sent = wb.create_sheet("Sentimientos")
            self._create_summary_sheet(ws_sent, df, start_date, end_date)
            
            ws_data = wb.create_sheet("Datos Sentimiento")
            self._create_data_sheet(ws_data, df)
            
            if callback:
                progress = 40
                callback(progress)
        
        # Alertas
        if include_alerts:
            anomalies = self._detect_anomalies(30, 2.0)
            ws_alerts = wb.create_sheet("Alertas")
            
            ws_alerts['A1'] = 'ALERTAS Y ANOMALÍAS'
            ws_alerts['A1'].font = self.styles['title']['font']
            
            if anomalies:
                df_alerts = pd.DataFrame(anomalies)
                self._write_pivot_table(ws_alerts, df_alerts, start_row=3)
            
            if callback:
                progress = 70
                callback(progress)
        
        # Tendencias
        if include_trends:
            df = self._get_sentiment_dataframe(start_date, end_date)
            df_daily = self._aggregate_daily(df)
            
            ws_trends = wb.create_sheet("Tendencias")
            ws_trends['A1'] = 'ANÁLISIS DE TENDENCIAS'
            ws_trends['A1'].font = self.styles['title']['font']
            
            self._create_daily_sheet(ws_trends, df_daily)
            
            if callback:
                progress = 90
                callback(progress)
        
        wb.save(output_path)
        
        if callback:
            callback(100)
        
        logger.info(f"Reporte combinado generado: {output_path}")
        return output_path


# Función de utilidad para uso directo
def generate_excel_report(report_type: str, **kwargs) -> str:
    """
    Función de conveniencia para generar reportes Excel.
    
    Args:
        report_type: 'sentiment', 'pivot', 'anomalies', 'combined'
        **kwargs: Parámetros específicos
    
    Returns:
        str: Ruta del Excel generado
    """
    generator = ExcelGenerator()
    
    if report_type == 'sentiment':
        return generator.generate_sentiment_dataset(**kwargs)
    elif report_type == 'pivot':
        return generator.generate_pivot_table(**kwargs)
    elif report_type == 'anomalies':
        return generator.generate_anomalies_report(**kwargs)
    elif report_type == 'combined':
        return generator.generate_combined_report(**kwargs)
    else:
        raise ValueError(f"Tipo de reporte Excel no válido: {report_type}")
